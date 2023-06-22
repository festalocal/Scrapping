# Importation des bibliothèques nécessaires
import pandas as pd
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, colors
from datetime import datetime

def format_time(time_str, new_format="%I:%M %p"):
    """
    Cette fonction convertit une chaîne de temps ISO 8601 en une nouvelle chaîne de temps.

    :param time_str: La chaîne de temps ISO 8601 à convertir.
    :type time_str: str
    :param new_format: Le format de la nouvelle chaîne de temps.
    :type new_format: str
    :return: La nouvelle chaîne de temps.
    :rtype: str
    """
    try:
        # Remplace "Z" par "+00:00" pour permettre la conversion en format datetime
        # Convertit en format datetime
        # Formate en chaîne de temps dans le nouveau format
        return datetime.fromisoformat(time_str.replace("Z", "+00:00")).strftime(new_format)
    except ValueError:
        return time_str

def adjust_columns_width(worksheet):
    """
    Cette fonction ajuste la largeur des colonnes d'une feuille de calcul Excel en fonction de la longueur des valeurs.

    :param worksheet: La feuille de calcul Excel à ajuster.
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :return: None
    """
    # Parcours de chaque colonne dans la feuille de calcul
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        # Parcours de chaque cellule dans la colonne
        for cell in column:
            try:
                # Si la longueur de la valeur de la cellule est plus grande que la longueur maximale, mise à jour de la longueur maximale
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        # Ajustement de la largeur de la colonne en fonction de la longueur maximale
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

def get_nested_value(obj, keys, default=''):
    """
    Cette fonction obtient une valeur imbriquée d'un objet en utilisant une liste de clés.

    :param obj: L'objet d'où obtenir la valeur.
    :type obj: dict
    :param keys: La liste des clés pour accéder à la valeur.
    :type keys: list of str
    :param default: La valeur par défaut à retourner si une clé n'est pas trouvée.
    :type default: any
    :return: La valeur obtenue.
    :rtype: any
    """
    for key in keys:
        if obj is None:
            return default
        if key in obj:
            obj = obj[key]
        else:
            return default
    return obj

def convert_json_to_excel(json_file, excel_file):
    """
    Cette fonction convertit les données JSON en un fichier Excel.

    :param json_file: Le fichier JSON à convertir.
    :type json_file: str
    :param excel_file: Le nom du fichier Excel de sortie.
    :type excel_file: str
    :return: None
    """
    # Ouverture du fichier JSON et chargement des données
    with open(json_file, encoding='utf-8') as f:
        data = json.load(f)

    events = []

    # Extraction des événements des données
    for obj in data:
        if 'events' in obj:
            events.extend(obj['events'])

    # Définition des en-têtes et de la largeur des colonnes pour le fichier Excel
    headers = [
        ('Titre', 30),
        ('Description', 50),
        ('Dates', 20),
        ('Heure de début', 20),
        ('Heure de fin', 20),
        ('Mode de participation', 30),
        ('Lieu', 30),
        ('Adresse', 30),
        ('Ville', 20),
        ('Latitude', 20),
        ('Longitude', 20),
        ('Mots-clés', 30),
        ('Agenda d\'origine', 30),
        ('Image', 30)
    ]

    # Initialisation des colonnes pour le fichier Excel
    columns = {header[0]: [] for header in headers}

    # Extraction des données des événements et ajout aux colonnes
    for event in events:
        columns['Titre'].append(get_nested_value(event, ['title', 'fr']))
        columns['Description'].append(get_nested_value(event, ['description', 'fr']))
        columns['Dates'].append(get_nested_value(event, ['dateRange', 'fr']))
        columns['Heure de début'].append(format_time(get_nested_value(event, ['lastTiming', 'begin'])))
        columns['Heure de fin'].append(format_time(get_nested_value(event, ['lastTiming', 'end'])))
        columns['Mode de participation'].append(get_nested_value(event, ['attendanceMode']))
        columns['Lieu'].append(get_nested_value(event, ['location', 'name']))
        columns['Adresse'].append(get_nested_value(event, ['location', 'address']))
        columns['Ville'].append(get_nested_value(event, ['location', 'city']))
        columns['Latitude'].append(get_nested_value(event, ['location', 'latitude']))
        columns['Longitude'].append(get_nested_value(event, ['location', 'longitude']))
        
        keywords = get_nested_value(event, ['keywords', 'fr'], [])
        if keywords is not None:
            keywords = [keyword for keyword in keywords if keyword is not None]
            columns['Mots-clés'].append(', '.join(keywords))
        else:
            columns['Mots-clés'].append('')
        
        columns['Agenda d\'origine'].append(get_nested_value(event, ['originAgenda', 'title']))
        image = get_nested_value(event, ['image', 'filename'])
        if image is not None:
            image = 'https://cibul.s3.amazonaws.com/' + image
        columns['Image'].append(image)

    # Création d'un DataFrame à partir des colonnes
    df = pd.DataFrame(columns)

    # Création de nouvelles colonnes en minuscules pour identifier les doublons
    df['Titre_lower'] = df['Titre'].str.lower()
    df['Dates_lower'] = df['Dates'].str.lower()
    df['Description_lower'] = df['Description'].str.lower()

    # Suppression des doublons
    df.drop_duplicates(subset=['Titre_lower', 'Dates_lower', 'Description_lower'], inplace=True)

    # Suppression des nouvelles colonnes en minuscules car elles ne sont plus nécessaires
    df.drop(columns=['Titre_lower', 'Dates_lower', 'Description_lower'], inplace=True)

    # Remplacement des valeurs vides par NaN dans la colonne 'Titre'
    df['Titre'].replace('', pd.np.nan, inplace=True)
    # Suppression des lignes où 'Titre' est NaN
    df.dropna(subset=['Titre'], inplace=True)

    # Réinitialisation de l'index du DataFrame
    df.reset_index(drop=True, inplace=True)

    # Création d'un nouveau classeur Excel et d'une nouvelle feuille de calcul
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Événements'
    
    # Ajout des en-têtes et mise en forme des cellules d'en-tête
    for idx, header in enumerate(headers):
        column_letter = chr(65 + idx)
        column_width = header[1]
        column_name = header[0]
        
        worksheet[f'{column_letter}1'] = column_name
        worksheet.column_dimensions[column_letter].width = column_width
        
        header_cell = worksheet[f'{column_letter}1']
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color='483D8B', end_color='483D8B', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajout des données et mise en forme des cellules de données
    for row_idx, row in df.iterrows():
        for idx, header in enumerate(headers):
            column_letter = chr(65 + idx)
            column_name = header[0]
            value = row[column_name]
            
            cell = worksheet[f'{column_letter}{row_idx + 2}']
            cell.value = value

            # Alignement du texte
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Style des cellules
            if row_idx % 2 == 0:  # Pour les lignes impaires
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            else:  # Pour les lignes paires
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Ajustement de la largeur des colonnes
    adjust_columns_width(worksheet)
    # Enregistrement du classeur Excel
    workbook.save(excel_file)

# Lancement de la fonction principale si ce fichier est exécuté comme script principal
if __name__ == '__main__':
    convert_json_to_excel("events.json", "evenements.xlsx")

import pandas as pd
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, colors
from datetime import datetime

def format_time(time_str, new_format="%I:%M %p"):
    try:
        return datetime.fromisoformat(time_str.replace("Z", "+00:00")).strftime(new_format)
    except ValueError:
        return time_str



def adjust_columns_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

def get_nested_value(obj, keys, default=''):
    for key in keys:
        if obj is None:
            return default
        if key in obj:
            obj = obj[key]
        else:
            return default
    return obj

def convert_json_to_excel(json_file, excel_file):
    with open(json_file, encoding='utf-8') as f:
        data = json.load(f)

    events = []
    
    for obj in data:
        if 'events' in obj:
            events.extend(obj['events'])

    headers = [
        ('Titre', 30),
        ('Description', 50),
        ('Dates', 20),
        ('Heure de début', 20),
        ('Heure de fin', 20),
        ('Mode de participation', 30),
        ('Lieu', 30),
        ('Adresse', 30),
        ('Ville', 20),
        ('Latitude', 20),
        ('Longitude', 20),
        ('Mots-clés', 30),
        ('Agenda d\'origine', 30),
        ('Image', 30)
    ]

    columns = {header[0]: [] for header in headers}
    
    for event in events:
        columns['Titre'].append(get_nested_value(event, ['title', 'fr']))
        columns['Description'].append(get_nested_value(event, ['description', 'fr']))
        columns['Dates'].append(get_nested_value(event, ['dateRange', 'fr']))
        columns['Heure de début'].append(format_time(get_nested_value(event, ['lastTiming', 'begin'])))
        columns['Heure de fin'].append(format_time(get_nested_value(event, ['lastTiming', 'end'])))
        columns['Mode de participation'].append(get_nested_value(event, ['attendanceMode']))
        columns['Lieu'].append(get_nested_value(event, ['location', 'name']))
        columns['Adresse'].append(get_nested_value(event, ['location', 'address']))
        columns['Ville'].append(get_nested_value(event, ['location', 'city']))
        columns['Latitude'].append(get_nested_value(event, ['location', 'latitude']))
        columns['Longitude'].append(get_nested_value(event, ['location', 'longitude']))
        
        keywords = get_nested_value(event, ['keywords', 'fr'], [])
        if keywords is not None:
            keywords = [keyword for keyword in keywords if keyword is not None]
            columns['Mots-clés'].append(', '.join(keywords))
        else:
            columns['Mots-clés'].append('')
        
        columns['Agenda d\'origine'].append(get_nested_value(event, ['originAgenda', 'title']))
        image = get_nested_value(event, ['image', 'filename'])
        if image is not None:
            image = 'https://cibul.s3.amazonaws.com/' + image
        columns['Image'].append(image)

    df = pd.DataFrame(columns)

    df['Titre_lower'] = df['Titre'].str.lower()
    df['Dates_lower'] = df['Dates'].str.lower()
    df['Description_lower'] = df['Description'].str.lower()

    df.drop_duplicates(subset=['Titre_lower', 'Dates_lower', 'Description_lower'], inplace=True)

    df.drop(columns=['Titre_lower', 'Dates_lower', 'Description_lower'], inplace=True)

    df['Titre'].replace('', pd.np.nan, inplace=True)
    df.dropna(subset=['Titre'], inplace=True)

    df.reset_index(drop=True, inplace=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Événements'
    
    for idx, header in enumerate(headers):
        column_letter = chr(65 + idx)
        column_width = header[1]
        column_name = header[0]
        
        worksheet[f'{column_letter}1'] = column_name
        worksheet.column_dimensions[column_letter].width = column_width
        
        header_cell = worksheet[f'{column_letter}1']
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color='483D8B', end_color='483D8B', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in df.iterrows():
        for idx, header in enumerate(headers):
            column_letter = chr(65 + idx)
            column_name = header[0]
            value = row[column_name]
            
            cell = worksheet[f'{column_letter}{row_idx + 2}']
            cell.value = value

            # Alignement du texte
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Style des cellules
            if row_idx % 2 == 0:  # Pour les lignes impaires
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            else:  # Pour les lignes paires
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    adjust_columns_width(worksheet)
    workbook.save(excel_file)

if __name__ == '__main__':
    convert_json_to_excel("events.json", "evenements.xlsx")
    